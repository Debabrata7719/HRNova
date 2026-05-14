"""
Employee Management Router
POST /api/employees            — Add a single employee (HR only)
POST /api/employees/bulk       — Bulk add from Excel, returns full summary (HR only)
POST /api/employees/bulk/stream — Bulk add with real-time SSE progress (HR only)
"""

import re
import io
import json
import random
import bcrypt
import yagmail
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from api.dependencies.auth import get_current_user
from src.tools.db_connection import get_db
from src.config import get_settings
from src.logger import get_logger

router = APIRouter()
logger = get_logger(__name__)

VALID_ROLES = {"EMPLOYEE", "HR"}
EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class AddEmployeeRequest(BaseModel):
    name: str
    email: str
    department: str
    role: str = "EMPLOYEE"


def generate_password(name: str) -> str:
    """
    Generate password: First4LettersOfCleanName + @ + 4 random digits
    "raj kumar" -> "Rajk@4821"
    "Ali"       -> "Ali@3392"
    """
    clean_name = "".join(name.split()).title()
    prefix = clean_name[:4] if len(clean_name) >= 4 else clean_name
    digits = str(random.randint(1000, 9999))
    return f"{prefix}@{digits}"


def send_welcome_email(employee_name: str, employee_email: str, password: str) -> bool:
    """
    Send welcome email with credentials.
    Called ONLY after successful DB insertion.
    Returns True on success, False on failure (non-fatal).
    """
    cfg = get_settings()

    if not cfg.EMAIL_ADDRESS or not cfg.EMAIL_APP_PASSWORD:
        logger.warning("[EMPLOYEE] Email not configured — skipping welcome email")
        return False

    subject = "Welcome to NovaHR — Your Account is Ready"
    body = f"""Hello {employee_name},

Your NovaHR account has been created by HR.

Your login credentials:
  Email    : {employee_email}
  Password : {password}

Please log in and change your password after your first login.
This is a system generated email, don't reply to it.

— NovaHR HR Team"""

    try:
        yag = yagmail.SMTP(
            user=cfg.EMAIL_ADDRESS,
            password=cfg.EMAIL_APP_PASSWORD,
        )
        yag.send(to=employee_email, subject=subject, contents=body)
        logger.info("[EMPLOYEE] Welcome email sent to %s", employee_email)
        return True
    except Exception as e:
        logger.error("[EMPLOYEE] Failed to send welcome email to %s: %s", employee_email, e)
        return False


@router.post("/employees")
def add_employee(request: AddEmployeeRequest, user=Depends(get_current_user)):
    """
    Add a new employee. HR only.

    Order of operations:
    1. Validate inputs
    2. Check email not duplicate
    3. Generate password
    4. Hash password
    5. INSERT into DB  — if this fails, raise error, NO email sent
    6. Send welcome email (best-effort, after successful insert only)
    7. Return employee data + generated password
    """
    if user.get("role") != "HR":
        raise HTTPException(status_code=403, detail="Only HR can add employees")

    role = request.role.upper()
    if role not in ("EMPLOYEE", "HR"):
        raise HTTPException(status_code=400, detail="Role must be EMPLOYEE or HR")

    name = request.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")

    email = request.email.strip().lower()
    if "@" not in email:
        raise HTTPException(status_code=400, detail="Invalid email address")

    department = request.department.strip()
    if not department:
        raise HTTPException(status_code=400, detail="Department is required")

    db = get_db()

    # Check duplicate
    existing = db.execute_query(
        "SELECT id FROM employees WHERE email = %s", (email,)
    )
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Email '{email}' is already registered"
        )

    # Generate + hash password
    password = generate_password(name)
    hashed = bcrypt.hashpw(
        password.encode("utf-8"),
        bcrypt.gensalt(rounds=10)
    ).decode("utf-8")

    # DB insertion — email NOT sent before this
    new_id = db.insert_query(
        """
        INSERT INTO employees (name, email, department, password, auth_role)
        VALUES (%s, %s, %s, %s, %s)
        """,
        (name, email, department, hashed, role)
    )

    if new_id is None:
        raise HTTPException(
            status_code=500,
            detail="Failed to insert employee into database"
        )

    logger.info("[EMPLOYEE] Added: %s (%s) id=%s", name, email, new_id)

    # Send welcome email ONLY after successful DB insert
    email_sent = send_welcome_email(name, email, password)

    return {
        "success": True,
        "employee": {
            "id": new_id,
            "name": name,
            "email": email,
            "department": department,
            "role": role,
        },
        "generated_password": password,
        "email_sent": email_sent,
        "message": (
            f"Employee '{name}' added successfully. "
            + ("Credentials emailed to the employee."
               if email_sent
               else "Email could not be sent — share the password manually.")
        )
    }


@router.post("/employees/bulk")
async def bulk_add_employees(
    file: UploadFile = File(...),
    user=Depends(get_current_user)
):
    """
    Bulk add employees from an Excel (.xlsx) file. HR only.

    Expected Excel columns (row 1 = header, data from row 2):
        name | email | department | role

    Validation per row:
    - name: required, non-empty
    - email: required, valid format, not duplicate in DB, not duplicate within file
    - department: required, non-empty
    - role: must be EMPLOYEE or HR (case-insensitive), defaults to EMPLOYEE if empty

    Order of operations per valid row:
    1. Insert into DB
    2. Send welcome email ONLY after successful insert

    Returns a full upload summary with success count, failure count, and per-row errors.
    """
    if user.get("role") != "HR":
        raise HTTPException(status_code=403, detail="Only HR can bulk add employees")

    # Validate file type
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    # Read file into memory
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file is empty or has no data rows")

    # Skip header row
    data_rows = rows[1:]

    db = get_db()

    # Fetch all existing emails from DB once (for duplicate check)
    existing = db.execute_query("SELECT email FROM employees") or []
    existing_emails = {row["email"].lower() for row in existing}

    results = []
    success_count = 0
    fail_count = 0
    seen_emails_in_file = set()  # track duplicates within the uploaded file

    for idx, row in enumerate(data_rows, start=2):  # start=2 because row 1 is header
        # Extract columns — handle None and extra whitespace
        name = str(row[0]).strip() if row[0] is not None else ""
        email = str(row[1]).strip().lower() if row[1] is not None else ""
        department = str(row[2]).strip() if row[2] is not None else ""
        role_raw = str(row[3]).strip().upper() if row[3] is not None else ""
        role = role_raw if role_raw in VALID_ROLES else "EMPLOYEE"

        # ── Validation ────────────────────────────────────────────────────────
        errors = []

        if not name or name.lower() == "none":
            errors.append("Name is required")

        if not email or email.lower() == "none":
            errors.append("Email is required")
        elif not EMAIL_REGEX.match(email):
            errors.append(f"Invalid email format: '{email}'")
        elif email in existing_emails:
            errors.append(f"Email already registered in database")
        elif email in seen_emails_in_file:
            errors.append(f"Duplicate email in uploaded file")

        if not department or department.lower() == "none":
            errors.append("Department is required")

        if role_raw and role_raw not in VALID_ROLES:
            errors.append(f"Invalid role '{role_raw}' — must be EMPLOYEE or HR")

        if errors:
            fail_count += 1
            results.append({
                "row": idx,
                "name": name or "(empty)",
                "email": email or "(empty)",
                "status": "failed",
                "errors": errors,
            })
            continue

        # Mark email as seen in this file
        seen_emails_in_file.add(email)

        # ── Generate password ─────────────────────────────────────────────────
        password = generate_password(name)
        hashed = bcrypt.hashpw(
            password.encode("utf-8"),
            bcrypt.gensalt(rounds=10)
        ).decode("utf-8")

        # ── DB insertion ──────────────────────────────────────────────────────
        new_id = db.insert_query(
            """
            INSERT INTO employees (name, email, department, password, auth_role)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (name, email, department, hashed, role)
        )

        if new_id is None:
            fail_count += 1
            results.append({
                "row": idx,
                "name": name,
                "email": email,
                "status": "failed",
                "errors": ["Database insertion failed"],
            })
            continue

        # Add to existing_emails so subsequent rows in same file detect it as duplicate
        existing_emails.add(email)
        logger.info("[BULK] Added employee: %s (%s)", name, email)

        # ── Send welcome email ONLY after successful DB insert ─────────────────
        email_sent = send_welcome_email(name, email, password)

        success_count += 1
        results.append({
            "row": idx,
            "name": name,
            "email": email,
            "status": "success",
            "email_sent": email_sent,
        })

    # Build summary
    failed_rows = [r for r in results if r["status"] == "failed"]
    success_rows = [r for r in results if r["status"] == "success"]

    logger.info(
        "[BULK] Upload complete — %d success, %d failed out of %d rows",
        success_count, fail_count, len(data_rows)
    )

    return {
        "total_rows": len(data_rows),
        "success_count": success_count,
        "fail_count": fail_count,
        "successful": success_rows,
        "failed": failed_rows,
    }


@router.post("/employees/bulk/stream")
async def bulk_add_employees_stream(
    file: UploadFile = File(...),
    user=Depends(get_current_user)
):
    """
    Bulk add employees from Excel with real-time SSE progress streaming.
    HR only.

    Streams JSON events as each row is processed:
      {"type": "progress", "row": 3, "status": "success", "name": "Raj", "email": "raj@co.com"}
      {"type": "progress", "row": 4, "status": "failed",  "name": "...", "email": "...", "errors": [...]}
      {"type": "done", "total_rows": 10, "success_count": 8, "fail_count": 2, "failed": [...]}
    """
    if user.get("role") != "HR":
        raise HTTPException(status_code=403, detail="Only HR can bulk add employees")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    contents = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file is empty or has no data rows")

    data_rows = rows[1:]

    async def event_generator():
        db = get_db()
        existing = db.execute_query("SELECT email FROM employees") or []
        existing_emails = {row["email"].lower() for row in existing}
        seen_in_file = set()
        failed_rows = []
        success_count = 0
        fail_count = 0

        for idx, row in enumerate(data_rows, start=2):
            name = str(row[0]).strip() if row[0] is not None else ""
            email = str(row[1]).strip().lower() if row[1] is not None else ""
            department = str(row[2]).strip() if row[2] is not None else ""
            role_raw = str(row[3]).strip().upper() if row[3] is not None else ""
            role = role_raw if role_raw in VALID_ROLES else "EMPLOYEE"

            errors = []
            if not name or name.lower() == "none":
                errors.append("Name is required")
            if not email or email.lower() == "none":
                errors.append("Email is required")
            elif not EMAIL_REGEX.match(email):
                errors.append(f"Invalid email format")
            elif email in existing_emails:
                errors.append("Email already registered")
            elif email in seen_in_file:
                errors.append("Duplicate email in file")
            if not department or department.lower() == "none":
                errors.append("Department is required")
            if role_raw and role_raw not in VALID_ROLES:
                errors.append(f"Invalid role '{role_raw}'")

            if errors:
                fail_count += 1
                entry = {"row": idx, "name": name or "(empty)", "email": email or "(empty)", "errors": errors}
                failed_rows.append(entry)
                event = {"type": "progress", "row": idx, "status": "failed",
                         "name": name or "(empty)", "email": email or "(empty)", "errors": errors}
                yield f"data: {json.dumps(event)}\n\n"
                continue

            seen_in_file.add(email)

            password = generate_password(name)
            hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=10)).decode("utf-8")

            new_id = db.insert_query(
                "INSERT INTO employees (name, email, department, password, auth_role) VALUES (%s, %s, %s, %s, %s)",
                (name, email, department, hashed, role)
            )

            if new_id is None:
                fail_count += 1
                entry = {"row": idx, "name": name, "email": email, "errors": ["Database insertion failed"]}
                failed_rows.append(entry)
                event = {"type": "progress", "row": idx, "status": "failed",
                         "name": name, "email": email, "errors": ["Database insertion failed"]}
                yield f"data: {json.dumps(event)}\n\n"
                continue

            existing_emails.add(email)
            email_sent = send_welcome_email(name, email, password)
            success_count += 1

            event = {"type": "progress", "row": idx, "status": "success",
                     "name": name, "email": email, "email_sent": email_sent}
            yield f"data: {json.dumps(event)}\n\n"

        # Final done event
        done_event = {
            "type": "done",
            "total_rows": len(data_rows),
            "success_count": success_count,
            "fail_count": fail_count,
            "failed": failed_rows,
        }
        yield f"data: {json.dumps(done_event)}\n\n"
        logger.info("[BULK STREAM] Done — %d success, %d failed", success_count, fail_count)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )
